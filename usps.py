'''
Here present command line options
usp marketing, or usp editorial

then ask for input

present options asking all titles with one usp

all titles with duplicate usps

generate report with total number of titles with each option

later create report of each title that has option

'''

# https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl import Workbook
import os
import sys
sys.path.append('C:\\Code\\uspChecks\\library')
import re
from usp_parser import uspHtmlParser
# from html_tags import tag_content
from sheet_info import headers, column_number
from sheet_data import get_sheetdata
import time 

startTime = time.time()

data = get_sheetdata("dataset/uspDataset_current.xlsx")

# for row in data.iter_rows('A{}:A{}'.format(2, 10)):
# 	for cell in row:
# 		print(cell.value)


# cmd line input: this file [0] 'title' [1] usp field [2]
# get output file name from input

file = sys.argv[0]
file_parts = os.path.splitext(file)
file_stem = file_parts[0]

# provide options for standard input based on headers dict

headers = headers(data)
input1 = sys.argv[1]
column_in_one = headers[input1]
col_num_in_one = column_number(column_in_one)
input2 = sys.argv[2]
column_in_two = headers[input2]
col_num_in_two = column_number(column_in_two)
isbn = headers['isbn']
isbn_col = column_number(isbn)
prod_cat = headers['product_category']
prod_cat_col = column_number(prod_cat)
market_us = headers['marketing_class_us']
market_us_col = column_number(market_us)
market_row = headers['marketing_class_row']
market_row_col = column_number(market_row)
market_dach = headers['marketing_class_dach']
market_dach_col = column_number(market_dach)



# number of rows
count = data.max_row

# add one for range operator
count += 1

buk = Workbook()

overview = buk.active
overview.title = 'usp html'

# more than one usp of particular type, input 2 from cmd line
# one for german, the other for english
usp_input2_DE = buk.create_sheet(input2+'_DE')
usp_input2_EN = buk.create_sheet(input2+'_EN')

# name the output headers
# make this dynamic
n_row_overview = 1
overview.cell(row=n_row_overview, column=1, value=input1)
overview.cell(row=n_row_overview, column=2, value="ISBN")
overview.cell(row=n_row_overview, column=3, value=input2)
overview.cell(row=n_row_overview, column=4, value="# HTML Tags")
overview.cell(row=n_row_overview, column=5, value="HTML Tags")
overview.cell(row=n_row_overview, column=6, value="Product Category")
overview.cell(row=n_row_overview, column=7, value="Marketing US")
overview.cell(row=n_row_overview, column=8, value="Marketing ROW")
overview.cell(row=n_row_overview, column=9, value="Marketing DACH")



n_row_input2 = 1
# german
n_row_de = 1 # headers
usp_input2_DE.cell(row=n_row_de, column=1, value=input1)
usp_input2_DE.cell(row=n_row_de, column=2, value="ISBN")
usp_input2_DE.cell(row=n_row_de, column=3, value=input2)
usp_input2_DE.cell(row=n_row_de, column=4, value="# of HTML tags")
usp_input2_DE.cell(row=n_row_de, column=5, value="# of HTML-USPs")
usp_input2_DE.cell(row=n_row_de, column=6, value="USP content")
usp_input2_DE.cell(row=n_row_de, column=7, value="Word Count")
usp_input2_DE.cell(row=n_row_de, column=8, value="Less than 3 Words")
usp_input2_DE.cell(row=n_row_de, column=9, value="More than 15 Words")
usp_input2_DE.cell(row=n_row_de, column=10, value="O words")
usp_input2_DE.cell(row=n_row_de, column=11, value="1 lowercase word")
usp_input2_DE.cell(row=n_row_de, column=12, value="the lowercase word")

# english
n_row_en = 1 # headers
usp_input2_EN.cell(row=n_row_en, column=1, value=input1)
usp_input2_EN.cell(row=n_row_en, column=2, value="ISBN")
usp_input2_EN.cell(row=n_row_en, column=3, value=input2)
usp_input2_EN.cell(row=n_row_en, column=4, value="# of HTML tags")
usp_input2_EN.cell(row=n_row_en, column=5, value="# of HTML-USPs")
usp_input2_EN.cell(row=n_row_en, column=6, value="USP content")
usp_input2_EN.cell(row=n_row_en, column=7, value="Word Count")
usp_input2_EN.cell(row=n_row_en, column=8, value="Less than 3 Words")
usp_input2_EN.cell(row=n_row_en, column=9, value="More than 15 Words")
usp_input2_EN.cell(row=n_row_en, column=10, value="O words")
usp_input2_EN.cell(row=n_row_en, column=11, value="1 lowercase word")
usp_input2_EN.cell(row=n_row_en, column=12, value="the lowercase word")



parser = uspHtmlParser()

for n in range(2, count):
	project_prelim = data.cell(row=n, column=16).value
	p = re.compile(r'\bpreliminary\b')
	if not p.search(project_prelim):
		n_row_overview += 1		
		inOne = data.cell(row=n, column=col_num_in_one).value
		inTwo = data.cell(row=n, column=col_num_in_two).value
		# print('overview inTwo', end='')
		# print(inTwo)		
		parser.feed(inTwo)
		html_total = len(parser.tags)
		overview.cell(row=n_row_overview, column=1, value=inOne)
		overview.cell(row=n_row_overview, column=2, value=data.cell(row=n, 
			column=isbn_col).value) # ISBN
		overview.cell(row=n_row_overview, column=3, value=inTwo)
		overview.cell(row=n_row_overview, column=4, 
			value=html_total)
		overview.cell(row=n_row_overview, column=5, 
			value=', '.join(parser.tags))
		overview.cell(row=n_row_overview, column=6, value=data.cell(row=n, 
			column=9).value)
		overview.cell(row=n_row_overview, column=7, value=data.cell(row=n, 
			column=market_us_col).value)
		overview.cell(row=n_row_overview, column=8, value=data.cell(row=n, 
			column=market_row_col).value)
		overview.cell(row=n_row_overview, column=9, value=data.cell(row=n, 
			column=market_dach_col).value)

		
		if html_total >= 1:
			langs = data.cell(row=n, column=4).value
			main_lang = langs.split()[0]

			if main_lang == 'EN' or main_lang == 'DE':	
				usps_parsed, one_word_lowercase = parser.usps_parsed()
				n_usps = len(usps_parsed)			
				# n_row_input2 += 1				

				if main_lang == 'EN':					
					n_row_en += 1		
					usp_input2_EN.cell(row=n_row_en, column=1, value=inOne)
					usp_input2_EN.cell(row=n_row_en, column=2, value=data.cell(row=n,
						column=isbn_col).value) # ISBN
					usp_input2_EN.cell(row=n_row_en, column=3, value=inTwo)
					usp_input2_EN.cell(row=n_row_en, column=4, value=html_total)
					usp_input2_EN.cell(row=n_row_en, column=5, value=n_usps)			
					usp_content = parser.usps_as_string(usps_parsed)
					usp_input2_EN.cell(row=n_row_en, column=6, value=usp_content)
					word_count = parser.word_count_summary(usps_parsed)
					usp_input2_EN.cell(row=n_row_en, column=7, value=', '.join(str(x) for x in word_count))
					flag = False
					for i in word_count:
						if i < 3:
							flag = True
							break
					usp_input2_EN.cell(row=n_row_en, column=8, value=flag)
					flag = False
					for i in word_count:
						if i > 15:
							flag = True
							break
					usp_input2_EN.cell(row=n_row_en, column=9, value=flag)
					flag = False
					for i in word_count:
						if i == 0:
							flag = True
							break
					usp_input2_EN.cell(row=n_row_en, column=10, value=flag)
					flag = False
					if one_word_lowercase:
						flag = True						
						usp_input2_EN.cell(row=n_row_en, column=12, value=str(one_word_lowercase))
					# for i in word_count:
					# 	if i == 1:
					# 		flag = True
					# 		break
					usp_input2_EN.cell(row=n_row_en, column=11, value=flag)

				if main_lang == 'DE':
					n_row_de += 1
					usp_input2_DE.cell(row=n_row_de, column=1, value=inOne)
					usp_input2_DE.cell(row=n_row_de, column=2, value=data.cell(row=n,
						column=isbn_col).value) # ISBN
					usp_input2_DE.cell(row=n_row_de, column=3, value=inTwo)
					usp_input2_DE.cell(row=n_row_de, column=4, value=html_total)
					usp_input2_DE.cell(row=n_row_de, column=5, value=n_usps)			
					usp_content = parser.usps_as_string(usps_parsed)
					usp_input2_DE.cell(row=n_row_de, column=6, value=usp_content)
					word_count = parser.word_count_summary(usps_parsed)
					usp_input2_DE.cell(row=n_row_de, column=7, value=', '.join(str(x) for x in word_count))
					flag = False
					for i in word_count:
						if i < 3:
							flag = True
							break
					usp_input2_DE.cell(row=n_row_de, column=8, value=flag)
					flag = False
					for i in word_count:
						if i > 15:
							flag = True
							break
					usp_input2_DE.cell(row=n_row_de, column=9, value=flag)
					flag = False
					for i in word_count:
						if i == 0:
							flag = True
							break
					usp_input2_DE.cell(row=n_row_de, column=10, value=flag)
					flag = False
					if one_word_lowercase:
						flag = True
						word = one_word_lowercase[0]
						usp_input2_DE.cell(row=n_row_de, column=12, value=str(one_word_lowercase))
					# for i in word_count:
					# 	if i == 1:
					# 		flag = True
					# 		break
					usp_input2_DE.cell(row=n_row_de, column=11, value=flag)

			

print_date = time.strftime("%d%m%y")
print_time = time.strftime("%I%M%S")

buk.save('results/'+file_stem+'_'+input2+'_'+print_date+'_'+print_time+'.xlsx')

print ('The script took {0} seconds !'.format(time.time() - startTime))