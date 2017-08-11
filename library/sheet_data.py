from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import sys
sys.path.append('C:\\Code\\uspChecks\\library')
from sheet_info import headers
import string

'''
function to get the data from Delilah file
'''
def get_sheetdata(file):

	wb = load_workbook(file)

	names = wb.get_sheet_names()

	return wb[names[0]] # data returned