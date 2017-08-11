'''
Here present command line options
usp marketing, or usp editorial

then ask for input

present options asking all titles with one usp

all titles with duplicate usps

generate report with total number of titles with each option

later create report of each title that has option

'''

# https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl import Workbook
import os
import sys
sys.path.append('C:\\Code\\uspChecks\\library')
import re
from html_tags import tag_content
from sheet_info import headers, column_number
from sheet_data import get_sheetdata
import time 

startTime = time.time()

data = get_sheetdata("dataset/uspDataset_test.xlsx")

# for row in data.iter_rows('A{}:A{}'.format(2, 10)):
# 	for cell in row:
# 		print(cell.value)

# get file name from input

file = sys.argv[0]
file_parts = os.path.splitext(file)
file_stem = file_parts[0]

# provide options for standard input based on headers dict

headers = headers(data)
input1 = sys.argv[1]
column_one = headers[input1]
col_num_one = column_number(column_one)
input2 = sys.argv[2]
column_two = headers[input2]
col_num_two = column_number(column_two)

# number of rows
count = data.max_row

buk = Workbook()

outsheet_full = buk.active
outsheet_full.title = 'usp total'

# create sheet for 1 usp
uspOne = buk.create_sheet('1 usp_editorial')

# name the output headers
# make this dynamic
n_row_full = 1
outsheet_full.cell(row=n_row_full, column=1, value=input1)
outsheet_full.cell(row=n_row_full, column=2, value="ISBN")
outsheet_full.cell(row=n_row_full, column=3, value=input2)
outsheet_full.cell(row=n_row_full, column=4, value="# USPs")


# 1 usp sheet
n_row_uspOne = 1
uspOne.cell(row=n_row_uspOne, column=1, value=input1)
uspOne.cell(row=n_row_uspOne, column=2, value="ISBN")
uspOne.cell(row=n_row_uspOne, column=3, value=input2)


for n in range(2, 10):
	project_prelim = data.cell(row=n, column=16).value
	p = re.compile(r'\bpreliminary\b')
	if not p.search(project_prelim):
		n_row_full += 1
		inOne = data.cell(row=n, column=col_num_one).value
		inTwo = data.cell(row=n, column=col_num_two).value
		usps = tag_content(inTwo)
		total = len(usps)
		outsheet_full.cell(row=n_row_full, column=1, value=inOne)		
		outsheet_full.cell(row=n_row_full, column=2, value=data.cell(row=n, 
			column=2).value) # ISBN
		outsheet_full.cell(row=n_row_full, column=3, value=inTwo)
		outsheet_full.cell(row=n_row_full, column=4, value=total)
		if total == 1:
			n_row_uspOne += 1
			uspOne.cell(row=n_row_uspOne, column=1, value=inOne)
			uspOne.cell(row=n_row_uspOne, column=2, value=data.cell(row=n,
				column=2).value) # ISBN
			uspOne.cell(row=n_row_uspOne, column=3, value=inTwo)

print_date = time.strftime("%d%m%y")
print_time = time.strftime("%I%M%S")

buk.save('results/'+file_stem+'_'+input2+'_'+print_date+'_'+print_time+'.xlsx')

print ('The script took {0} seconds !'.format(time.time() - startTime))