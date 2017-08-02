'''
Here present command line options
usp marketing, or usp editorial

then ask for input

present options asking all titles with one usp

all titles with duplicate usps

generate report with total number of titles with each option

later create report of each title that has option

'''

# https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl import Workbook
import os
import sys
sys.path.append('C:\\Code\\uspChecks\\library')
import re
from sheetInfo import headers, column_number
from sheetData import get_sheetdata
import time 

startTime = time.time()

data = get_sheetdata("dataset/uspDataset_current.xlsx")

# get file name from input

file = sys.argv[0]
file_parts = os.path.splitext(file)
file_stem = file_parts[0]

headers = headers(data)
input1 = sys.argv[1]
column_one = headers[input1]
col_num_one = column_number(column_one)
input2 = sys.argv[2]
column_two = headers[input2]
col_num_two = column_number(column_two)

# number of rows
count = data.max_row

buk = Workbook()

outsheet_full = buk.active
outsheet_full.title = 'usp total'

# create sheet for 1 usp
uspDup = buk.create_sheet('usp dups')

# name the output headers
# make this dynamic
n_row_full = 1
outsheet_full.cell(row=n_row_full, column=1, value=input1)
outsheet_full.cell(row=n_row_full, column=2, value="ISBN")
outsheet_full.cell(row=n_row_full, column=3, value=input2)
outsheet_full.cell(row=n_row_full, column=4, value="# USPs")

# 1 usp sheet
n_row_uspDup = 1
uspDup.cell(row=n_row_uspDup, column=1, value=input1)
uspDup.cell(row=n_row_uspDup, column=2, value=input2)

for n in range(2, count):
	project_prelim = data.cell(row=n, column=16).value
	p = re.compile(r'\bpreliminary\b')
	if not p.search(project_prelim):
		n_row_full += 1
		inOne = data.cell(row=n, column=col_num_one).value
		inTwo = data.cell(row=n, column=col_num_two).value
		p = re.compile(r'<p>')
		pm = p.findall(str(inTwo))
		if len(pm) is 0:
			div = re.compile(r'<div>')
			divm = div.findall(str(inTwo))
			total = len(divm)
		else:
			total = len(pm)		
		outsheet_full.cell(row=n_row_full, column=1, value=inOne)
		outsheet_full.cell(row=n_row_full, column=2, value=data.cell(row=n,
			column=2).value) # ISBN
		outsheet_full.cell(row=n_row_full, column=3, value=inTwo)
		outsheet_full.cell(row=n_row_full, column=4, value=total)
		if total > 1:
			pTag = re.compile(r'<p>(.*?)</p>')
			divTag = re.compile(r'<div>(.*?)</div>')
			if len(pm):
				p = re.compile(r'<p>(.*?)</p>')
				pm = p.findall(inTwo)
				diff = len(pm) - len(set(pm))
				# diff = pm					
			else: # in this case div tags are at least 2
				div = re.compile(r'<div>(.*?)</div>')
				divm = div.findall(inTwo)
				diff = len(divm) - len(set(divm))
				# diff = divm
			if diff:
				n_row_uspDup += 1
				uspDup.cell(row=n_row_uspDup, column=1, value=inOne)
				uspDup.cell(row=n_row_uspDup, column=2, value=data.cell(row=n,
					column=2).value)
				uspDup.cell(row=n_row_uspDup, column=3, value=inTwo)
				uspDup.cell(row=n_row_uspDup, column=4, value=diff)

buk.save('results/'+file_stem+'_'+input2+'.xlsx')

print ('The script took {0} seconds !'.format(time.time() - startTime))


# re.compile(r'<p>(.*?)</p>')
