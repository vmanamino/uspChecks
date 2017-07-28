from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import sys
sys.path.append('C:\\Code\\uspChecks\\library')
from specs import headers
import string

wb = load_workbook("dataset/uspDataset_current.xlsx")

names = wb.get_sheet_names()

data = wb[names[0]]

columns = headers(data)

print(columns['A'])

